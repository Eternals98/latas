from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from src.models import Cliente, EstadoVentaEnum, Pago, Venta


def _create_venta(
    db_session,
    *,
    referencia: str,
    creado_en: datetime,
    valor_total: str = "100.00",
    tipo: str = "formal",
    estado: str = EstadoVentaEnum.ACTIVO.value,
    descripcion: str = "Venta reporte",
    cliente: Cliente | None = None,
) -> Venta:
    venta = Venta(
        empresa="latas_sas",
        tipo=tipo,
        numero_referencia=referencia,
        descripcion=descripcion,
        fecha_venta=creado_en.date(),
        valor_total=Decimal(valor_total),
        cliente_id=cliente.id if cliente else None,
        estado=estado,
        creado_en=creado_en,
        modificado_en=creado_en,
    )
    db_session.add(venta)
    db_session.flush()
    db_session.add(Pago(venta_id=venta.id, medio="efectivo", monto=Decimal(valor_total)))
    db_session.commit()
    return venta


def _xlsx_rows(response):
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def test_monthly_listing_filters_period_and_excludes_annulled(client, db_session):
    cliente = Cliente(nombre="Cliente Uno", telefono="3001112233")
    db_session.add(cliente)
    db_session.commit()

    april = _create_venta(
        db_session,
        referencia="ABR-001",
        creado_en=datetime(2026, 4, 1, 0, 0, 0),
        cliente=cliente,
    )
    _create_venta(db_session, referencia="MAY-001", creado_en=datetime(2026, 5, 1, 0, 0, 0))
    _create_venta(
        db_session,
        referencia="ABR-ANULADA",
        creado_en=datetime(2026, 4, 15, 12, 0, 0),
        estado=EstadoVentaEnum.ANULADO.value,
    )

    response = client.get("/api/ventas?mes=4&anio=2026")

    assert response.status_code == 200
    body = response.json()
    assert body["mes"] == 4
    assert body["anio"] == 2026
    assert [item["id"] for item in body["items"]] == [april.id]
    assert body["items"][0]["cliente"]["nombre"] == "Cliente Uno"
    assert body["items"][0]["pagos"][0]["monto"] == "100.00"


def test_monthly_listing_empty_month_returns_empty_items(client, db_session):
    _create_venta(db_session, referencia="MAR-001", creado_en=datetime(2026, 3, 31, 23, 59, 59))

    response = client.get("/api/ventas?mes=4&anio=2026")

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["resumen_mensual"] == {
        "mes": 4,
        "anio": 2026,
        "cantidad_ventas": 0,
        "valor_total": "0.00",
    }


def test_monthly_listing_summary_matches_returned_rows(client, db_session):
    _create_venta(db_session, referencia="ABR-001", creado_en=datetime(2026, 4, 2), valor_total="125.50")
    _create_venta(db_session, referencia="ABR-002", creado_en=datetime(2026, 4, 3), valor_total="74.50")

    response = client.get("/api/ventas?mes=4&anio=2026")

    assert response.status_code == 200
    body = response.json()
    assert body["resumen_mensual"]["cantidad_ventas"] == 2
    assert body["resumen_mensual"]["valor_total"] == "200.00"


def test_export_formal_contains_only_active_formal_ventas(client, db_session):
    _create_venta(db_session, referencia="FORMAL", creado_en=datetime(2026, 4, 2), tipo="formal")
    _create_venta(db_session, referencia="INFORMAL", creado_en=datetime(2026, 4, 2), tipo="informal")
    _create_venta(
        db_session,
        referencia="FORMAL-ANULADA",
        creado_en=datetime(2026, 4, 2),
        tipo="formal",
        estado=EstadoVentaEnum.ANULADO.value,
    )

    response = client.get("/api/ventas/export?tipo=formal&mes=4&anio=2026")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    rows = _xlsx_rows(response)
    assert rows[0] == (
        "Fecha",
        "Empresa",
        "Tipo",
        "Numero referencia",
        "Cliente",
        "Telefono",
        "Descripcion",
        "Valor total",
        "Medios de pago",
        "Estado",
    )
    assert [row[3] for row in rows[1:]] == ["FORMAL"]


def test_export_informal_contains_only_active_informal_ventas(client, db_session):
    _create_venta(db_session, referencia="FORMAL", creado_en=datetime(2026, 4, 2), tipo="formal")
    _create_venta(db_session, referencia="INFORMAL", creado_en=datetime(2026, 4, 2), tipo="informal")

    response = client.get("/api/ventas/export?tipo=informal&mes=4&anio=2026")

    assert response.status_code == 200
    rows = _xlsx_rows(response)
    assert [row[3] for row in rows[1:]] == ["INFORMAL"]


def test_export_invalid_tipo_returns_400(client, db_session):
    response = client.get("/api/ventas/export?tipo=otro")

    assert response.status_code == 400
    assert "tipo" in response.json()["detail"]


def test_export_rejects_partial_period(client, db_session):
    response = client.get("/api/ventas/export?tipo=formal&mes=4")

    assert response.status_code == 400
    assert "mes y anio" in response.json()["detail"]


def test_export_no_matches_returns_header_only_workbook(client, db_session):
    response = client.get("/api/ventas/export?tipo=formal&mes=4&anio=2026")

    assert response.status_code == 200
    rows = _xlsx_rows(response)
    assert len(rows) == 1


def test_export_preserves_special_characters_and_long_descriptions(client, db_session):
    cliente = Cliente(nombre="Jose Alvarez & Cia", telefono="300")
    db_session.add(cliente)
    db_session.commit()
    descripcion = "Latas especiales - " + ("descripcion larga " * 8) + "ñ"
    _create_venta(
        db_session,
        referencia="ESP-001",
        creado_en=datetime(2026, 4, 2),
        descripcion=descripcion,
        cliente=cliente,
    )

    response = client.get("/api/ventas/export?tipo=formal&mes=4&anio=2026")

    rows = _xlsx_rows(response)
    assert rows[1][4] == "Jose Alvarez & Cia"
    assert rows[1][6] == descripcion

    persisted = db_session.execute(select(Venta).where(Venta.numero_referencia == "ESP-001")).scalar_one()
    assert persisted.descripcion == descripcion
