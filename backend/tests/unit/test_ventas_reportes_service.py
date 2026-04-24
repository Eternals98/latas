from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.models import Pago, Venta
from src.services.ventas_service import (
    VentaValidationError,
    build_monthly_summary,
    build_ventas_xlsx,
    month_range,
    validate_period,
)


def test_month_range_uses_start_inclusive_end_exclusive_dates():
    start, end = month_range(2, 2028)

    assert start == date(2028, 2, 1)
    assert end == date(2028, 3, 1)


@pytest.mark.parametrize(
    ("mes", "anio"),
    [(0, 2026), (13, 2026), (1, 1999)],
)
def test_validate_period_rejects_invalid_period(mes, anio):
    with pytest.raises(VentaValidationError):
        validate_period(mes, anio)


def test_build_monthly_summary_uses_decimal_totals():
    ventas = [
        Venta(
            empresa="latas_sas",
            tipo="formal",
            numero_referencia="A",
            descripcion="A",
            valor_total=Decimal("10.10"),
        ),
        Venta(
            empresa="latas_sas",
            tipo="formal",
            numero_referencia="B",
            descripcion="B",
            valor_total=Decimal("20.20"),
        ),
    ]

    summary = build_monthly_summary(ventas, mes=4, anio=2026)

    assert summary.cantidad_ventas == 2
    assert summary.valor_total == Decimal("30.30")


def test_build_ventas_xlsx_writes_headers_and_rows():
    venta = Venta(
        id=1,
        empresa="latas_sas",
        tipo="formal",
        numero_referencia="F-001",
        descripcion="Venta",
        fecha_venta=date(2026, 4, 2),
        valor_total=Decimal("50.00"),
        estado="activo",
        creado_en=datetime(2026, 4, 2, 9, 30, 0),
    )
    venta.pagos = [Pago(id=1, venta_id=1, medio="efectivo", monto=Decimal("50.00"))]

    payload = build_ventas_xlsx([venta])
    rows = list(load_workbook(BytesIO(payload)).active.iter_rows(values_only=True))

    assert rows[0][0:4] == ("Fecha", "Empresa", "Tipo", "Numero referencia")
    assert rows[1] == (
        "2026-04-02",
        "latas_sas",
        "formal",
        "F-001",
        None,
        None,
        "Venta",
        "50.00",
        "efectivo: 50.00",
        "activo",
    )
