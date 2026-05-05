from __future__ import annotations

from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.models.cash_event import CashEvent
from src.models.cash_movement import CashMovement
from src.models.cash_session import CashSession
from src.models.company import Company
from src.models.customer import Customer
from src.models.payment_method import PaymentMethod
from src.models.profile import Profile
from src.models.transaction import Transaction
from src.models.transaction_payment import TransactionPayment
from src.api.schemas.sales import to_money
from src.api.schemas.historic_migration import HistoricMigrationResponse


class HistoricMigrationError(Exception):
    pass


def _now() -> datetime:
    return datetime.now(UTC)


def _sheet_timestamp(sheet_date: date, *, hour: int, minute: int = 0, second: int = 0) -> datetime:
    return datetime.combine(sheet_date, time(hour=hour, minute=minute, second=second), tzinfo=UTC)


def _row_timestamp(sheet_date: date, sequence: int) -> datetime:
    base = _sheet_timestamp(sheet_date, hour=8, minute=0)
    return base + timedelta(minutes=sequence)


def _normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: object | None) -> str:
    return _normalize_text(value).lower()


def _as_decimal(value: object | None) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    try:
        return to_money(value)  # type: ignore[arg-type]
    except Exception:
        return Decimal("0.00")


def _is_method_marker(value: object | None, method_name: str) -> bool:
    text = _normalize_key(value)
    return bool(text) and text == _normalize_key(method_name)


def _resolve_company(db: Session, company_name: str) -> Company:
    cleaned = company_name.strip()
    if not cleaned:
        company = db.execute(select(Company).where(Company.name.ilike("Genérico"))).scalars().first()
        if company is None:
            company = db.execute(select(Company).where(Company.name.ilike("Generico"))).scalars().first()
        if company is None:
            company = Company(name="Genérico", is_active=True)
            db.add(company)
            db.flush()
        return company
    company = db.execute(select(Company).where(Company.name.ilike(cleaned))).scalars().first()
    if company is None:
        company = Company(name=cleaned, is_active=True)
        db.add(company)
        db.flush()
    return company


def _resolve_customer(db: Session, name: str | None) -> Customer:
    cleaned = _normalize_text(name)
    if not cleaned:
        customer = db.execute(select(Customer).where(Customer.is_generic.is_(True), Customer.is_active.is_(True))).scalars().first()
        if customer is None:
            raise HistoricMigrationError("No existe cliente genérico activo.")
        return customer
    customer = db.execute(select(Customer).where(Customer.name.ilike(cleaned))).scalars().first()
    if customer is None:
        customer = Customer(name=cleaned, is_generic=False, is_active=True)
        db.add(customer)
        db.flush()
    return customer


def _resolve_payment_method(db: Session, name: str) -> PaymentMethod:
    cleaned = name.strip()
    method = db.execute(select(PaymentMethod).where(PaymentMethod.name.ilike(cleaned))).scalars().first()
    if method is None:
        code = cleaned.upper().replace(" ", "_")
        method = PaymentMethod(name=cleaned, code=code, affects_cash=cleaned.upper() in {"EFECTIVO", "ENTREGA"}, is_active=True)
        db.add(method)
        db.flush()
    return method


def _sheet_is_consolidated(title: str) -> bool:
    return title.strip().lower() == "consolidado"


def _resolve_sheet_month(filename: str | None, month_override: int | None = None) -> int | None:
    if month_override is not None and 1 <= month_override <= 12:
        return month_override
    text = _normalize_key(filename)
    if "enero" in text:
        return 1
    if "febrero" in text:
        return 2
    if "marzo" in text:
        return 3
    if "abril" in text:
        return 4
    if "mayo" in text:
        return 5
    if "junio" in text:
        return 6
    if "julio" in text:
        return 7
    if "agosto" in text:
        return 8
    if "septiembre" in text or "setiembre" in text:
        return 9
    if "octubre" in text:
        return 10
    if "noviembre" in text:
        return 11
    if "diciembre" in text:
        return 12
    return None


def _payment_columns() -> list[tuple[int, str]]:
    return [
        (6, "EFECTIVO"),
        (7, "ENTREGA"),
        (8, "TARJETA LATAS"),
        (9, "TARJETA TOMAS"),
        (10, "BANCOLOMBIA LATAS"),
        (11, "BANCOLOMBIA TOMAS"),
        (12, "BBVA LATAS"),
        (13, "BBVA TOMAS"),
        (14, "GIRO"),
        (15, "DAVIVIENDA"),
        (16, "NEQUI TG"),
        (17, "NEQUI AB"),
        (18, "NEQUI OTRO"),
    ]


def _collect_row_payments(ws, row_idx: int, fallback_method_name: str, fallback_amount: Decimal) -> list[tuple[str, Decimal]]:
    payments: list[tuple[str, Decimal]] = []
    normalized_fallback_amount = abs(fallback_amount)
    normalized_fallback_method = _normalize_key(fallback_method_name)
    for column_index, method_name in _payment_columns():
        value = ws.cell(row_idx, column_index).value
        amount = _as_decimal(value)
        if amount != Decimal("0.00"):
            payments.append((method_name, amount))
            continue
        if _is_method_marker(value, method_name) and normalized_fallback_amount > Decimal("0.00"):
            payments.append((method_name, normalized_fallback_amount))

    if payments:
        return payments

    if normalized_fallback_method and normalized_fallback_amount > Decimal("0.00"):
        return [(fallback_method_name, normalized_fallback_amount)]

    return []


def _row_error_context(ws, row_idx: int) -> str:
    values = []
    for col_idx, label in _payment_columns():
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value not in (None, ""):
            values.append(f"{label}={cell_value}")
    company = _normalize_text(ws.cell(row_idx, 1).value)
    document_number = _normalize_text(ws.cell(row_idx, 2).value)
    description = _normalize_text(ws.cell(row_idx, 3).value)
    amount = ws.cell(row_idx, 4).value
    payment_method = _normalize_text(ws.cell(row_idx, 5).value)
    prefix = f"Hoja {ws.title} fila {row_idx}"
    parts = [prefix]
    if company:
        parts.append(f"empresa={company}")
    if document_number:
        parts.append(f"documento={document_number}")
    if description:
        parts.append(f"descripcion={description}")
    if amount not in (None, ""):
        parts.append(f"valor={amount}")
    if payment_method:
        parts.append(f"forma_pago={payment_method}")
    if values:
        parts.append("pagos=" + ", ".join(values))
    return " | ".join(parts)


def _create_cash_session(db: Session, *, sheet_date: date, actor: Profile) -> CashSession:
    session = CashSession(
        id=str(uuid4()),
        session_date=sheet_date,
        opening_cash=Decimal("0.00"),
        closing_cash_expected=None,
        closing_cash_counted=None,
        difference_amount=None,
        status="open",
        opened_by=actor.id,
        closed_by=None,
        opened_at=_sheet_timestamp(sheet_date, hour=8, minute=0),
        closed_at=None,
    )
    db.add(session)
    db.flush()
    db.add(
        CashEvent(
            id=str(uuid4()),
            cash_session_id=session.id,
            event_type="open",
            actor_id=actor.id,
            event_at=_sheet_timestamp(sheet_date, hour=8, minute=0),
            payload={"session_date": sheet_date.isoformat(), "source": "historic_migration"},
            note="Apertura histórica",
        )
    )
    return session


def migrate_historic_excel(
    db: Session,
    *,
    content: bytes,
    actor: Profile,
    filename: str | None = None,
    month: str | None = None,
) -> HistoricMigrationResponse:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HistoricMigrationError("El archivo no es un Excel válido.") from exc

    warnings: list[str] = []
    imported_rows = 0
    sale_transactions = 0
    cancelled_transactions = 0
    grouped_document_count = 0
    cash_movements = 0
    cash_in_movements = 0
    cash_out_movements = 0
    vault_in_movements = 0
    payment_rows = 0
    period_start: date | None = None
    period_end: date | None = None
    month_override = None
    if month is not None and str(month).strip():
        try:
            month_override = int(str(month).strip())
        except Exception:
            raise HistoricMigrationError("El mes seleccionado no es válido.")

    sheet_month = _resolve_sheet_month(filename, month_override)
    if sheet_month is None:
        warnings.append("No se pudo inferir el mes desde el nombre del archivo; se usará enero por defecto.")
        sheet_month = 1

    processed_sheet_index = 0
    for ws in wb.worksheets:
        if _sheet_is_consolidated(ws.title):
            continue
        processed_sheet_index += 1

        sheet_date = date(2026, sheet_month, min(processed_sheet_index, 28))
        try:
            sheet_day = int(str(ws.title).strip())
            sheet_date = date(2026, sheet_month, sheet_day)
        except Exception:
            warnings.append(f"La hoja {ws.title} no tiene nombre numérico; se usa fecha provisional {sheet_date.isoformat()}.")

        period_start = sheet_date if period_start is None or sheet_date < period_start else period_start
        period_end = sheet_date if period_end is None or sheet_date > period_end else period_end
        sheet_session = _create_cash_session(db, sheet_date=sheet_date, actor=actor)
        sheet_cash_in = Decimal("0.00")
        sheet_cash_out = Decimal("0.00")
        sheet_vault_in = Decimal("0.00")
        sheet_row_sequence = 0

        grouped_rows: dict[tuple[str, str], dict[str, object]] = {}
        standalone_rows: list[dict[str, object]] = []

        for row_idx in range(5, ws.max_row + 1):
            company_value = _normalize_text(ws.cell(row_idx, 1).value)
            document_number = _normalize_text(ws.cell(row_idx, 2).value)
            description = _normalize_text(ws.cell(row_idx, 3).value)
            amount = _as_decimal(ws.cell(row_idx, 4).value)
            payment_method_name = _normalize_text(ws.cell(row_idx, 5).value)
            row_payments = _collect_row_payments(ws, row_idx, payment_method_name, amount)

            if not any([company_value, document_number, description, amount, payment_method_name, row_payments]):
                continue

            if amount == Decimal("0.00") and not payment_method_name:
                continue

            imported_rows += 1
            sheet_row_sequence += 1
            row_timestamp = _row_timestamp(sheet_date, sheet_row_sequence)

            row_payload = {
                "row_idx": row_idx,
                "sequence": sheet_row_sequence,
                "row_timestamp": row_timestamp,
                "company_value": company_value,
                "document_number": document_number,
                "description": description,
                "amount": amount,
                "payment_method_name": payment_method_name,
                "row_payments": row_payments,
            }

            if document_number:
                company_key = company_value or _normalize_text(ws.cell(max(row_idx - 1, 1), 1).value)
                group_key = (company_key.lower(), document_number.lower())
                grouped_rows.setdefault(
                    group_key,
                    {
                        "company_value": company_value,
                        "document_number": document_number,
                        "description_parts": [],
                        "row_payments": [],
                        "cash_lines": [],
                        "is_cancelled": False,
                        "rows": [],
                        "row_total": Decimal("0.00"),
                        "first_timestamp": row_timestamp,
                    },
                )
                group = grouped_rows[group_key]
                if company_value:
                    group["company_value"] = company_value
                if description:
                    group["description_parts"].append(description)
                group["row_payments"].extend(row_payments)
                group["rows"].append(row_payload)
                group["first_timestamp"] = min(group["first_timestamp"], row_timestamp)  # type: ignore[arg-type]
                group["row_total"] = to_money(
                    Decimal(str(group["row_total"])) + sum((abs(payment_amount) for _, payment_amount in row_payments), start=Decimal("0.00"))
                )
                group["is_cancelled"] = bool(group["is_cancelled"]) or amount < 0 or "devol" in description.lower() or "anul" in description.lower()
            elif payment_method_name.upper() == "ENTREGA":
                standalone_rows.append(row_payload)
            else:
                standalone_rows.append(row_payload)

        def emit_transaction(
            *,
            company_value: str,
            document_number: str | None,
            description: str,
            row_payments: list[tuple[str, Decimal]],
            is_cancelled: bool,
            transaction_amount: Decimal,
            row_count: int,
            transaction_at: datetime,
        ) -> None:
            nonlocal sale_transactions, cancelled_transactions, payment_rows, cash_movements, cash_in_movements, cash_out_movements, vault_in_movements, sheet_cash_in, sheet_cash_out, sheet_vault_in
            if not company_value and not document_number and not description:
                return
            company = _resolve_company(db, company_value or "Genérico")
            customer = _resolve_customer(db, None)
            transaction_id = str(uuid4())
            created_at = _now()
            normalized_amount = abs(transaction_amount)
            transaction = Transaction(
                id=transaction_id,
                company_id=company.id,
                customer_id=customer.id,
                cash_session_id=None,
                transaction_date=transaction_at,
                document_type="other",
                document_number=document_number or None,
                description=description or "Migración histórica",
                transaction_type="sale",
                total_amount=normalized_amount,
                status="confirmed",
                payment_difference_amount=Decimal("0.00"),
                payment_difference_reason=None,
                created_by=actor.id,
                updated_by=actor.id,
                cancelled_by=None,
                created_at=created_at,
                updated_at=created_at,
                cancelled_at=None,
                cancellation_reason=None,
            )
            db.add(transaction)
            db.flush()
            if row_payments:
                for method_name, payment_amount in row_payments:
                    payment_amount = abs(payment_amount)
                    if payment_amount <= Decimal("0.00"):
                        continue
                    method = _resolve_payment_method(db, method_name)
                    db.add(
                        TransactionPayment(
                            id=str(uuid4()),
                            transaction_id=transaction_id,
                            payment_method_id=method.id,
                            amount=to_money(payment_amount),
                            created_at=created_at,
                        )
                    )
                    payment_rows += 1
            else:
                method = _resolve_payment_method(db, "EFECTIVO")
                db.add(
                    TransactionPayment(
                        id=str(uuid4()),
                        transaction_id=transaction_id,
                        payment_method_id=method.id,
                        amount=normalized_amount,
                        created_at=created_at,
                    )
                )
                payment_rows += 1

            cash_amount = sum((payment_amount for method_name, payment_amount in row_payments if _normalize_key(method_name) in {"efectivo", "cash"}), start=Decimal("0.00"))
            delivery_amount = sum((payment_amount for method_name, payment_amount in row_payments if _normalize_key(method_name) == "entrega"), start=Decimal("0.00"))

            if is_cancelled and cash_amount > 0:
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=transaction_id,
                        cash_session_id=sheet_session.id,
                        movement_date=transaction_at.date(),
                        movement_type="cash_out",
                        amount=to_money(cash_amount),
                        description=f"Anulación histórica {document_number or transaction_id}",
                        created_by=actor.id,
                        created_at=created_at,
                    )
                )
                cash_movements += 1
                cash_out_movements += 1
                sheet_cash_out += cash_amount
            elif not is_cancelled and cash_amount > 0:
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=transaction_id,
                        cash_session_id=sheet_session.id,
                        movement_date=transaction_at.date(),
                        movement_type="cash_in",
                        amount=to_money(cash_amount),
                        description=f"Ingreso histórico {document_number or transaction_id}",
                        created_by=actor.id,
                        created_at=created_at,
                    )
                )
                cash_movements += 1
                cash_in_movements += 1
                sheet_cash_in += cash_amount

            if delivery_amount > 0:
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=transaction_id,
                        cash_session_id=sheet_session.id,
                        movement_date=transaction_at.date(),
                        movement_type="cash_out",
                        amount=to_money(delivery_amount),
                        description=f"Entrega histórica {document_number or transaction_id}",
                        created_by=actor.id,
                        created_at=created_at,
                    )
                )
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=transaction_id,
                        cash_session_id=sheet_session.id,
                        movement_date=transaction_at.date(),
                        movement_type="vault_in",
                        amount=to_money(delivery_amount),
                        description=f"Entrega histórica a bóveda {document_number or transaction_id}",
                        created_by=actor.id,
                        created_at=created_at,
                    )
                )
                cash_movements += 2
                cash_out_movements += 1
                vault_in_movements += 1
                sheet_cash_out += delivery_amount
                sheet_vault_in += delivery_amount

            if is_cancelled:
                transaction.status = "cancelled"
                transaction.cancelled_by = actor.id
                transaction.cancelled_at = created_at
                transaction.cancellation_reason = "Migración histórica"

            sale_transactions += 1
            if is_cancelled:
                cancelled_transactions += 1

        for group in grouped_rows.values():
            company_value = str(group["company_value"])
            document_number = str(group["document_number"])
            description_parts = [part for part in group["description_parts"] if part]
            description = " | ".join(dict.fromkeys(description_parts)) if description_parts else "Migración histórica"
            row_payments = list(group["row_payments"])
            if len(row_payments) > 1:
                payment_names = {name for name, _ in row_payments}
                if "EFECTIVO" in payment_names and "ENTREGA" in payment_names:
                    warnings.append(
                        f"Hoja {ws.title} documento {document_number}: mezcla efectivo y entrega en varias filas."
                    )
            if len(group["rows"]) > 1:
                grouped_document_count += 1
                warnings.append(f"Hoja {ws.title} documento {document_number}: se consolidaron {len(group['rows'])} filas en una sola transacción.")
            tx_amount = sum((abs(amount) for _, amount in row_payments), start=Decimal("0.00"))
            if tx_amount <= Decimal("0.00"):
                tx_amount = Decimal(str(group["row_total"])) if Decimal(str(group["row_total"])) > 0 else Decimal("0.00")
            if tx_amount <= Decimal("0.00"):
                first_row_idx = int(group["rows"][0]["row_idx"]) if group["rows"] else 0
                raise HistoricMigrationError(
                    f"No se pudo determinar un valor válido para la transacción. {_row_error_context(ws, first_row_idx)}"
                )
            emit_transaction(
                company_value=company_value,
                document_number=document_number,
                description=description,
                row_payments=row_payments,
                is_cancelled=bool(group["is_cancelled"]),
                transaction_amount=tx_amount,
                row_count=len(group["rows"]),
                transaction_at=group["first_timestamp"],  # type: ignore[arg-type]
            )

        for row_payload in standalone_rows:
            company_value = str(row_payload["company_value"]) or ""
            document_number = str(row_payload["document_number"]) or ""
            description = str(row_payload["description"]) or "Migración histórica"
            amount = row_payload["amount"] if isinstance(row_payload["amount"], Decimal) else Decimal("0.00")
            payment_method_name = str(row_payload["payment_method_name"]) or ""
            row_payments = list(row_payload["row_payments"])
            description_lower = description.lower()
            explicit_devolution = "devolucion" in description_lower or "devolución" in description_lower or "dev" in description_lower
            is_cancelled = amount < 0 or "anul" in description_lower or explicit_devolution
            is_vault_entry = amount < 0 and not explicit_devolution
            if payment_method_name.upper() == "ENTREGA" and not document_number:
                if amount <= Decimal("0.00"):
                    raise HistoricMigrationError(
                        f"No se pudo determinar un valor válido para la entrega. {_row_error_context(ws, int(row_payload['row_idx']))}"
                    )
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=None,
                        cash_session_id=sheet_session.id,
                        movement_date=sheet_date,
                        movement_type="cash_out",
                        amount=abs(amount),
                        description=description or "Entrega histórica",
                        created_by=actor.id,
                        created_at=_now(),
                    )
                )
                db.add(
                    CashMovement(
                        id=str(uuid4()),
                        transaction_id=None,
                        cash_session_id=sheet_session.id,
                        movement_date=sheet_date,
                        movement_type="vault_in",
                        amount=abs(amount),
                        description=description or "Entrega histórica a bóveda",
                        created_by=actor.id,
                        created_at=_now(),
                    )
                )
                cash_movements += 2
                sheet_cash_out += abs(amount)
                sheet_vault_in += abs(amount)
                continue
            if is_vault_entry:
                row_payments = [("ENTREGA", abs(amount))]
                payment_method_name = "ENTREGA"
                is_cancelled = False
                warnings.append(
                    f"Hoja {ws.title} fila {row_payload['row_idx']}: valor negativo tratado como ENTREGA para bóveda."
                )
            if not row_payments and amount <= Decimal("0.00"):
                raise HistoricMigrationError(
                    f"No se pudo determinar un valor válido para la transacción. {_row_error_context(ws, int(row_payload['row_idx']))}"
                )
            emit_transaction(
                company_value=company_value or "Genérico",
                document_number=document_number or None,
                description=description,
                row_payments=row_payments,
                is_cancelled=is_cancelled,
                transaction_amount=abs(sum((abs(payment_amount) for _, payment_amount in row_payments), start=Decimal("0.00")) or amount),
                row_count=1,
                transaction_at=row_payload["row_timestamp"],  # type: ignore[arg-type]
            )

        closing_amount = to_money(sheet_cash_in - sheet_cash_out)
        sheet_session.status = "closed"
        sheet_session.closing_cash_expected = closing_amount
        sheet_session.closing_cash_counted = closing_amount
        sheet_session.difference_amount = Decimal("0.00")
        sheet_session.closed_by = actor.id
        sheet_session.closed_at = _sheet_timestamp(sheet_date, hour=18, minute=0)
        db.add(
            CashEvent(
                id=str(uuid4()),
                cash_session_id=sheet_session.id,
                event_type="close",
                actor_id=actor.id,
                event_at=_sheet_timestamp(sheet_date, hour=18, minute=0),
                payload={
                    "session_date": sheet_date.isoformat(),
                    "expected_cash": f"{closing_amount:.2f}",
                    "counted_cash": f"{closing_amount:.2f}",
                    "difference_amount": "0.00",
                    "source": "historic_migration",
                },
                note="Cierre histórico",
            )
        )

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise

    return HistoricMigrationResponse(
        imported_sheets=sum(1 for ws in wb.worksheets if not _sheet_is_consolidated(ws.title)),
        imported_rows=imported_rows,
        sale_transactions=sale_transactions,
        cancelled_transactions=cancelled_transactions,
        grouped_document_count=grouped_document_count,
        cash_movements=cash_movements,
        cash_in_movements=cash_in_movements,
        cash_out_movements=cash_out_movements,
        vault_in_movements=vault_in_movements,
        payment_rows=payment_rows,
        warnings=warnings,
        period_start=period_start,
        period_end=period_end,
    )
