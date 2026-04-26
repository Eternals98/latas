from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
import re
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.api.schemas.ventas import CreatePagoRequest, CreateVentaRequest
from src.models import Cliente, EmpresaEnum, TipoVentaEnum
from src.services.ventas_service import VentaValidationError, create_venta_with_pagos, to_money


class ImportVentasError(Exception):
    pass


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFD", text)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def _contains_ref_error(value: object) -> bool:
    return "#ref!" in _normalize_text(value)


def _parse_money(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        if value == 0:
            return None
        return to_money(value)
    if isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
        if decimal_value == 0:
            return None
        return to_money(decimal_value)

    text = str(value).strip()
    if not text or text == "-":
        return None
    if _contains_ref_error(text):
        return None

    cleaned = text.replace("$", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    elif "." in cleaned:
        parts = cleaned.split(".")
        if len(parts) > 2:
            cleaned = "".join(parts)
        elif len(parts) == 2 and len(parts[-1]) == 3:
            cleaned = "".join(parts)

    cleaned = re.sub(r"[^0-9\.-]", "", cleaned)
    if not cleaned:
        return None

    try:
        amount = Decimal(cleaned)
    except Exception:
        return None
    if amount == 0:
        return None
    return to_money(amount)


def _map_empresa(raw_value: object) -> str:
    text = _normalize_text(raw_value)
    if not text:
        return EmpresaEnum.GENERICO.value
    if "tomas" in text:
        return EmpresaEnum.TOMAS_GOMEZ.value
    if "latas" in text:
        return EmpresaEnum.LATAS_SAS.value
    return EmpresaEnum.GENERICO.value


def _map_tipo(raw_value: object, default_tipo: str) -> str:
    text = _normalize_text(raw_value)
    if "formal" in text:
        return TipoVentaEnum.FORMAL.value
    if "informal" in text:
        return TipoVentaEnum.INFORMAL.value
    return default_tipo


def _map_medio(raw_value: object) -> str | None:
    text = _normalize_text(raw_value)
    if not text:
        return None
    if "bancolombia" in text and "tomas" in text:
        return "bancolombia_tomas"
    if "bancolombia" in text and "latas" in text:
        return "bancolombia_latas"
    if "bbva" in text and "tomas" in text:
        return "bbva_tomas"
    if "bbva" in text and "latas" in text:
        return "bbva_latas"
    if "tarjeta" in text and "tomas" in text:
        return "tarjeta_tomas"
    if "tarjeta" in text and "latas" in text:
        return "tarjeta_latas"
    if "davivienda" in text or "daviplata" in text:
        return "davivienda"
    if "nequi" in text:
        return "nequi"
    if "efectivo" in text:
        return "efectivo"
    if "entrega" in text:
        return "entrega"
    if "otro" in text:
        return "otro"
    if "banco" in text:
        return "otro"
    return None


HEADER_ALIASES = {
    "empresa": "empresa",
    "tipo": "tipo",
    "descripcion": "descripcion",
    "valor": "valor",
    "forma pago": "forma_pago",
    "forma de pago": "forma_pago",
    "metodo de pago": "forma_pago",
    "metodo pago": "forma_pago",
    "efectivo": "efectivo",
    "entrega": "entrega",
    "tarjeta latas": "tarjeta_latas",
    "tarjeta tomas": "tarjeta_tomas",
    "bancolombia latas": "bancolombia_latas",
    "bancolombia tomas": "bancolombia_tomas",
    "bbva latas": "bbva_latas",
    "bbva tomas": "bbva_tomas",
    "davivienda": "davivienda",
    "nequi": "nequi",
    "otros": "otro",
    "otro": "otro",
}

PAYMENT_COLUMNS = (
    "efectivo",
    "entrega",
    "tarjeta_latas",
    "tarjeta_tomas",
    "bancolombia_latas",
    "bancolombia_tomas",
    "bbva_latas",
    "bbva_tomas",
    "davivienda",
    "nequi",
    "otro",
)

GENERIC_CLIENT_NAME = "CLIENTE GENERICO"
GENERIC_CLIENT_PHONE = "0000000000"
GENERIC_CLIENT_NORMALIZED = "cliente generico"


def _get_or_create_generic_cliente(db: Session) -> Cliente:
    existing = (
        db.execute(select(Cliente).where(Cliente.nombre_normalizado == GENERIC_CLIENT_NORMALIZED))
        .scalars()
        .one_or_none()
    )
    if existing is not None:
        if existing.telefono != GENERIC_CLIENT_PHONE:
            existing.telefono = GENERIC_CLIENT_PHONE
            db.commit()
            db.refresh(existing)
        return existing

    cliente = Cliente(
        nombre=GENERIC_CLIENT_NAME,
        telefono=GENERIC_CLIENT_PHONE,
        nombre_normalizado=GENERIC_CLIENT_NORMALIZED,
        estado="activo",
    )
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


def _find_header_row(sheet) -> tuple[int, dict[str, int]] | None:
    header_row = 4
    if sheet.max_row < header_row:
        return None
    values = [sheet.cell(row=header_row, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    mapped: dict[str, int] = {}
    for col_idx, raw in enumerate(values, start=1):
        normalized = _normalize_text(raw)
        if not normalized:
            continue
        if normalized in HEADER_ALIASES:
            mapped[HEADER_ALIASES[normalized]] = col_idx
    if "descripcion" in mapped and "valor" in mapped:
        return header_row, mapped
    return None


def import_ventas_excel(
    db: Session,
    *,
    file_bytes: bytes,
    mes: int,
    anio: int,
    default_tipo: str,
) -> dict:
    if default_tipo not in (TipoVentaEnum.FORMAL.value, TipoVentaEnum.INFORMAL.value):
        raise ImportVentasError("tipo_default debe ser formal o informal.")

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes))
    except Exception as exc:
        raise ImportVentasError("Archivo Excel invalido o no soportado.") from exc

    created_count = 0
    skipped_count = 0
    processed_sheets = 0
    skipped_sheets: list[str] = []
    errors: list[str] = []
    generic_cliente = _get_or_create_generic_cliente(db)

    for sheet in workbook.worksheets:
        sheet_name = (sheet.title or "").strip()
        if not sheet_name.isdigit():
            skipped_sheets.append(sheet_name or "<sin_nombre>")
            continue

        day = int(sheet_name)
        try:
            fecha_venta = date(anio, mes, day)
        except ValueError:
            skipped_sheets.append(sheet_name)
            continue

        header_found = _find_header_row(sheet)
        if header_found is None:
            skipped_sheets.append(sheet_name)
            continue

        processed_sheets += 1
        header_row, columns = header_found
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
            if all(value in (None, "", "-") for value in row_values):
                continue
            if any(_contains_ref_error(value) for value in row_values if value is not None):
                skipped_count += 1
                continue

            description = str(sheet.cell(row=row_idx, column=columns["descripcion"]).value or "").strip()
            if not description:
                skipped_count += 1
                continue

            empresa = _map_empresa(sheet.cell(row=row_idx, column=columns.get("empresa", 0)).value if "empresa" in columns else "")
            tipo = _map_tipo(
                sheet.cell(row=row_idx, column=columns.get("tipo", 0)).value if "tipo" in columns else "",
                default_tipo,
            )
            valor = _parse_money(sheet.cell(row=row_idx, column=columns["valor"]).value)

            pagos_payload: list[CreatePagoRequest] = []
            for payment_column in PAYMENT_COLUMNS:
                col_idx = columns.get(payment_column)
                if col_idx is None:
                    continue
                amount = _parse_money(sheet.cell(row=row_idx, column=col_idx).value)
                if amount is None:
                    continue
                try:
                    pagos_payload.append(CreatePagoRequest(medio=payment_column, monto=amount))
                except Exception:
                    continue

            if not pagos_payload:
                forma_col = columns.get("forma_pago")
                mapped_medio = _map_medio(sheet.cell(row=row_idx, column=forma_col).value if forma_col else "")
                if mapped_medio and valor is not None:
                    try:
                        pagos_payload.append(CreatePagoRequest(medio=mapped_medio, monto=valor))
                    except Exception:
                        pass

            if not pagos_payload:
                skipped_count += 1
                continue

            pagos_total = sum((pago.monto for pago in pagos_payload), start=Decimal("0.00"))
            valor_total = valor if valor is not None else pagos_total
            if valor_total != pagos_total:
                valor_total = pagos_total

            try:
                payload = CreateVentaRequest(
                    empresa=empresa,
                    tipo=tipo,
                    numero_referencia=f"IMPORT-{anio:04d}{mes:02d}-{day:02d}-{row_idx:04d}",
                    descripcion=description,
                    fecha_venta=fecha_venta,
                    valor_total=valor_total,
                    cliente_id=generic_cliente.id,
                    pagos=pagos_payload,
                )
                create_venta_with_pagos(db, payload)
                created_count += 1
            except (VentaValidationError, ValueError) as exc:
                skipped_count += 1
                if len(errors) < 30:
                    errors.append(f"Hoja {sheet_name}, fila {row_idx}: {exc}")
            except Exception as exc:  # noqa: BLE001
                skipped_count += 1
                if len(errors) < 30:
                    errors.append(f"Hoja {sheet_name}, fila {row_idx}: {exc}")

    return {
        "creadas": created_count,
        "omitidas": skipped_count,
        "hojas_procesadas": processed_sheets,
        "hojas_omitidas": skipped_sheets,
        "errores": errors,
    }
